from flask import Flask, render_template, request, send_file
import pandas as pd
import openpyxl
from werkzeug.utils import secure_filename
import os
import tempfile
from datetime import datetime

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'file' not in request.files:
            return 'No file part'
        file = request.files['file']
        if file.filename == '':
            return 'No selected file'
        if file:
            filename = secure_filename(file.filename)
            with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                file_path = temp_file.name
                file.save(file_path)

            df = pd.read_excel(file_path, header=1)
            global last_data
            last_data = df[['Name', 'Surname', 'E-mail', 'Market', 'Va a ser PCC?', 'B2E User Name']].tail(25)
            data = last_data.to_dict(orient='records')

            return render_template('base.html', data=data)
    
    return render_template('base.html', data=None)

@app.route('/process', methods=['POST'])
def process():
    selected_rows = request.form.getlist('rows')
    selected_indices = [int(row) for row in selected_rows]

    if any(idx < 0 or idx >= len(last_data) for idx in selected_indices):
        return "Índice seleccionado está fuera de rango.", 400

    plantilla = 'PlantillaSTEP4.xlsx'
    wb = openpyxl.load_workbook(plantilla)
    ws = wb.active

    destination_row = 7
    for idx in selected_indices:
        name = last_data.iloc[idx]['Name']
        surname = last_data.iloc[idx]['Surname']
        email = last_data.iloc[idx]['E-mail']
        market = last_data.iloc[idx]['Market']
        pcc_status = last_data.iloc[idx]['Va a ser PCC?']
        b2e_username = last_data.iloc[idx]['B2E User Name']

        ws[f'C{destination_row}'] = name
        ws[f'D{destination_row}'] = surname
        ws[f'E{destination_row}'] = email

        if pcc_status == 'Y':
            if market == 'DACH':
                ws[f'F{destination_row}'] = "/+4940210918145 /+43122709858 /+41445295828"
            elif market == 'France':
                ws[f'F{destination_row}'] = "/+33180037979"
            elif market == 'Spain':
                ws[f'F{destination_row}'] = "/+34932952130"
            elif market == 'Italy':
                ws[f'F{destination_row}'] = "/+390109997099"
        else:
            ws[f'F{destination_row}'] = ""

        ws[f'Q{destination_row}'] = email
        ws[f'R{destination_row}'] = b2e_username

        destination_row += 1

    now = datetime.now()
    formatted_date = now.strftime("%Y%m%d_%H%M")
    output_file = os.path.join(tempfile.gettempdir(), f'D365_STEP4_CCH_{formatted_date}.xlsx')
    wb.save(output_file)

    return send_file(output_file, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
